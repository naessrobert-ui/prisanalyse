#!/usr/bin/env python3
"""
Portfolio rebalancer for Nordea-style multi-sheet portfolio workbooks.

Features
--------
- Reads one workbook with one portfolio per sheet.
- Classifies holdings into user-defined categories.
- Builds a reference allocation from adjustable reference rules.
- Calculates active bets vs reference.
- Supports either:
  * go fully to reference ok
  * reference + user-defined active bets
  * absolute target category weights
- Proposes category trades and routed fund-level trades.
- Skips trades smaller than a configurable minimum threshold.
- Writes a detailed Excel workbook with the results.
- Can generate a starter config based on the uploaded workbook.

Typical use
-----------
1) Create a starter config
   python portfolio_rebalancer.py init-config \
       --input "2025.04.07 Vekt PBPM NO_3G.xlsx" \
       --config-out rebalancer_config.json

2) Edit the config: categories, reference rules, active bets, minimum size.

3) Run the rebalance
   python portfolio_rebalancer.py run \
       --input "2025.04.07 Vekt PBPM NO_3G.xlsx" \
       --config rebalancer_config.json \
       --output rebalance_output.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


KNOWN_DEFAULT_CATEGORIES: Dict[str, str] = {
    "Nordea PB Norsk Aksje Portefolje B Acc NOK": "equity_norway",
    "Nordea 1 - Global Portfolio Fund BF-NOK": "equity_global_developed",
    "Nordea 2 - BetaPlus Enhan Global Eq Fd BF-NOK": "equity_global_developed",
    "Nordea Discretionary Global Equity C Acc NOK": "equity_global_developed",
    "Nordea Emerging Market Equities Fund C Acc NOK": "equity_global_em",
    "Nordea FRN Pensjon C Acc NOK": "fi_norway_short",
    "Nordea Kort Obligasjon Pluss C Acc NOK": "fi_norway_short",
    "Nordea Obligasjon III C Acc NOK": "fi_norway_long",
    "Nordea 1 - European Corporate Bond Fund HBCN-NOK": "fi_global",
    "Nordea 1 - Global High Yield Bond Fund HBCN-NOK": "fi_global",
    "Nordea 1 - US Corporate Bond Fund HBC-NOK": "fi_global",
    "Nordea Allokeringsfond Fund C Acc NOK": "allocation",
}

CATEGORY_DESCRIPTIONS: Dict[str, str] = {
    "cash": "Kontanter",
    "allocation": "Blandet / allokeringsfond",
    "equity_norway": "Aksjer Norge",
    "equity_global_developed": "Aksjer utland (utviklede markeder)",
    "equity_global_em": "Aksjer utland (vekstmarkeder)",
    "fi_norway_short": "Renter Norge korte",
    "fi_norway_long": "Renter Norge lange",
    "fi_global": "Renter utland",
}

REFERENCE_CATEGORIES: List[str] = [
    "cash",
    "allocation",
    "equity_norway",
    "equity_global_developed",
    "equity_global_em",
    "fi_norway_short",
    "fi_norway_long",
    "fi_global",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


class ConfigError(Exception):
    """Raised when the user config is inconsistent."""


class DataError(Exception):
    """Raised when the workbook cannot be parsed as expected."""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace("%", "")
    if text == "":
        return 0.0
    return float(text)


def parse_equity_share_from_portfolio_name(name: str) -> float:
    normalized = normalize_text(name)
    if normalized.upper() == "EG":
        return 1.0
    match = re.search(r"G(\d{2})(\d{2})", normalized)
    if not match:
        raise ConfigError(
            f"Klarte ikke å tolke aksjeandel fra porteføljenavn '{name}'. "
            "Bruk en override i config['reference']['portfolio_equity_share_overrides']."
        )
    return int(match.group(1)) / 100.0


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def read_workbook_holdings(input_path: Path, exposure_column: str) -> pd.DataFrame:
    workbook = load_workbook(input_path, data_only=True)
    records: List[Dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]
        header_map = {normalize_text(value): idx for idx, value in enumerate(headers)}

        required_headers = ["Portfolio", "Security name", "MV", exposure_column]
        missing = [h for h in required_headers if normalize_text(h) not in header_map]
        if missing:
            raise DataError(
                f"Fanen '{sheet_name}' mangler kolonner: {missing}. "
                f"Fant disse kolonnene: {[normalize_text(x) for x in headers if x is not None]}"
            )

        portfolio_idx = header_map["Portfolio"]
        security_idx = header_map["Security name"]
        mv_idx = header_map["MV"]
        expo_idx = header_map[normalize_text(exposure_column)]
        model_idx = header_map.get("Model portfolio")
        country_idx = header_map.get("Country")

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            portfolio = normalize_text(row[portfolio_idx])
            security = row[security_idx]
            mv = safe_float(row[mv_idx])
            expo = row[expo_idx]
            model_portfolio = normalize_text(row[model_idx]) if model_idx is not None else ""
            country = normalize_text(row[country_idx]) if country_idx is not None else ""

            if expo is None:
                continue

            security_name = normalize_text(security)
            if not security_name:
                continue

            # Keep only concrete holding rows, not subtotal rows.
            # In this workbook, the real cash row has a security name starting with 'PB Norway:'
            # and holding rows have a model portfolio or a recognizable security name.
            is_cash_position = security_name.startswith("PB Norway:")
            is_security_holding = bool(model_portfolio)
            if not (is_cash_position or is_security_holding):
                continue

            records.append(
                {
                    "sheet_name": sheet_name,
                    "portfolio": sheet_name,
                    "internal_portfolio_name": portfolio,
                    "row_number": row_number,
                    "security_name": security_name,
                    "country": country,
                    "model_portfolio": model_portfolio,
                    "mv": mv,
                    "exposure_pct": safe_float(expo),
                }
            )

    if not records:
        raise DataError("Fant ingen beholdningsrader i arbeidsboken.")

    df = pd.DataFrame(records)
    return df


def make_default_config(input_path: Path, exposure_column: str = "Lev. expo. distr. (PF)") -> Dict[str, Any]:
    holdings = read_workbook_holdings(input_path, exposure_column=exposure_column)
    unique_securities = sorted(holdings["security_name"].dropna().unique().tolist())

    security_categories: Dict[str, str] = {}
    for security in unique_securities:
        if security.startswith("PB Norway:"):
            security_categories[security] = "cash"
        else:
            security_categories[security] = KNOWN_DEFAULT_CATEGORIES.get(security, "UNMAPPED")

    portfolios = sorted(holdings["portfolio"].unique().tolist())

    default_buy_security_by_category = {
        "equity_norway": "Nordea PB Norsk Aksje Portefolje B Acc NOK",
        "equity_global_developed": "Nordea 1 - Global Portfolio Fund BF-NOK",
        "equity_global_em": "Nordea Emerging Market Equities Fund C Acc NOK",
        "fi_norway_short": "Nordea FRN Pensjon C Acc NOK",
        "fi_norway_long": "Nordea Obligasjon III C Acc NOK",
        "fi_global": "Nordea 1 - European Corporate Bond Fund HBCN-NOK",
        "allocation": "Nordea Allokeringsfond Fund C Acc NOK",
        "cash": next((s for s in unique_securities if s.startswith("PB Norway:")), "CASH"),
    }

    return {
        "exposure_column": exposure_column,
        "minimum_trade_pct": 0.09,
        "round_decimals": 4,
        "security_categories": security_categories,
        "reference": {
            "derive_equity_share_from_sheet_name": True,
            "portfolio_equity_share_overrides": {},
            "equity_norway_within_equity": 0.15,
            "em_within_international_equity": 0.15,
            "cash_target_pct": 0.0,
            "allocation_target_pct": 0.0,
            "fi_norway_within_fi": None,
            "fi_long_within_norway_fi": None,
        },
        "active_bets": {
            "mode": "go_to_reference",
            "residual_category": {
                "__default__": "fi_global",
                "EG": "equity_global_developed",
            },
            "category_active_bets_pct": {portfolio: {} for portfolio in portfolios},
            "absolute_target_weights_pct": {},
        },
        "trade_routing": {
            "prefer_existing_largest_holding": True,
            "default_buy_security_by_category": default_buy_security_by_category,
            "portfolio_security_overrides": {},
        },
    }


def classify_holdings(holdings: pd.DataFrame, config: Dict[str, Any]) -> pd.DataFrame:
    category_map = config.get("security_categories", {})
    classified = holdings.copy()

    def classify_security(security_name: str) -> str:
        if security_name not in category_map:
            return "UNMAPPED"
        return category_map[security_name]

    classified["category"] = classified["security_name"].map(classify_security)
    unmapped = classified[classified["category"].isin(["", "UNMAPPED"])]["security_name"].unique().tolist()
    if unmapped:
        raise ConfigError(
            "Følgende verdipapirer mangler kategori i config['security_categories']: "
            + ", ".join(sorted(unmapped))
        )
    return classified


def summarize_current_weights(classified: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        classified.groupby(["portfolio", "category"], as_index=False)
        .agg(current_weight_pct=("exposure_pct", "sum"), current_mv=("mv", "sum"))
    )
    return grouped


def portfolio_total_mv(classified_portfolio: pd.DataFrame) -> float:
    return float(classified_portfolio["mv"].sum())


def current_category_map(current_weights: pd.DataFrame, portfolio: str) -> Dict[str, float]:
    subset = current_weights[current_weights["portfolio"] == portfolio]
    mapping = {row["category"]: float(row["current_weight_pct"]) for _, row in subset.iterrows()}
    return {category: mapping.get(category, 0.0) for category in REFERENCE_CATEGORIES}


def _validate_share(name: str, value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    if not (0.0 <= float(value) <= 1.0):
        raise ConfigError(f"'{name}' må være mellom 0 og 1. Fikk {value}.")
    return float(value)


def build_reference_weights_for_portfolio(
    portfolio: str,
    current_by_category: Dict[str, float],
    config: Dict[str, Any],
) -> Dict[str, float]:
    reference = config["reference"]
    overrides = reference.get("portfolio_equity_share_overrides", {})

    if portfolio in overrides:
        equity_share = float(overrides[portfolio])
    elif reference.get("derive_equity_share_from_sheet_name", True):
        equity_share = parse_equity_share_from_portfolio_name(portfolio)
    else:
        raise ConfigError(
            f"Ingen aksjereferanse funnet for {portfolio}. Sett override i config['reference']['portfolio_equity_share_overrides']."
        )

    equity_norway_within_equity = _validate_share(
        "equity_norway_within_equity", reference.get("equity_norway_within_equity", 0.15)
    )
    em_within_international_equity = _validate_share(
        "em_within_international_equity", reference.get("em_within_international_equity", 0.15)
    )
    cash_target_pct = float(reference.get("cash_target_pct", 0.0))
    allocation_target_pct = float(reference.get("allocation_target_pct", 0.0))
    fi_norway_within_fi = _validate_share("fi_norway_within_fi", reference.get("fi_norway_within_fi"))
    fi_long_within_norway_fi = _validate_share(
        "fi_long_within_norway_fi", reference.get("fi_long_within_norway_fi")
    )

    if equity_share < 0 or equity_share > 1:
        raise ConfigError(f"Aksjeandel for {portfolio} må være mellom 0 og 1. Fikk {equity_share}.")

    target: Dict[str, float] = {category: 0.0 for category in REFERENCE_CATEGORIES}
    target["cash"] = cash_target_pct * 100.0
    target["allocation"] = allocation_target_pct * 100.0

    equity_total_pct = equity_share * 100.0
    fi_total_pct = 100.0 - target["cash"] - target["allocation"] - equity_total_pct
    if fi_total_pct < -1e-8:
        raise ConfigError(
            f"Referansen for {portfolio} summerer til mer enn 100 %. "
            f"Kontanter + allokering + aksjer = {100.0 - fi_total_pct:.4f} %."
        )
    fi_total_pct = max(fi_total_pct, 0.0)

    target["equity_norway"] = equity_total_pct * equity_norway_within_equity
    international_equity_pct = equity_total_pct - target["equity_norway"]
    target["equity_global_em"] = international_equity_pct * em_within_international_equity
    target["equity_global_developed"] = international_equity_pct - target["equity_global_em"]

    current_fi_total = (
        current_by_category.get("fi_norway_short", 0.0)
        + current_by_category.get("fi_norway_long", 0.0)
        + current_by_category.get("fi_global", 0.0)
    )
    current_norway_fi = current_by_category.get("fi_norway_short", 0.0) + current_by_category.get("fi_norway_long", 0.0)
    current_long_norway_fi = current_by_category.get("fi_norway_long", 0.0)

    if fi_total_pct == 0:
        target["fi_norway_short"] = 0.0
        target["fi_norway_long"] = 0.0
        target["fi_global"] = 0.0
    else:
        if fi_norway_within_fi is None:
            if current_fi_total > 0:
                fi_norway_within_fi = current_norway_fi / current_fi_total
            else:
                fi_norway_within_fi = 0.5
        fi_norway_pct = fi_total_pct * fi_norway_within_fi
        fi_global_pct = fi_total_pct - fi_norway_pct

        if fi_long_within_norway_fi is None:
            if current_norway_fi > 0:
                fi_long_within_norway_fi = current_long_norway_fi / current_norway_fi
            else:
                fi_long_within_norway_fi = 0.5
        fi_norway_long_pct = fi_norway_pct * fi_long_within_norway_fi
        fi_norway_short_pct = fi_norway_pct - fi_norway_long_pct

        target["fi_norway_short"] = fi_norway_short_pct
        target["fi_norway_long"] = fi_norway_long_pct
        target["fi_global"] = fi_global_pct

    total = sum(target.values())
    if not math.isclose(total, 100.0, rel_tol=0, abs_tol=1e-6):
        raise ConfigError(f"Intern feil: referansevekter for {portfolio} summerer til {total:.6f} og ikke 100.")
    return target


def get_residual_category(config: Dict[str, Any], portfolio: str) -> str:
    residual_cfg = config.get("active_bets", {}).get("residual_category", {})
    return residual_cfg.get(portfolio, residual_cfg.get("__default__", "fi_global"))


def build_target_weights(
    portfolio: str,
    reference_weights: Dict[str, float],
    config: Dict[str, Any],
) -> Tuple[Dict[str, float], List[str]]:
    active_cfg = config.get("active_bets", {})
    mode = active_cfg.get("mode", "go_to_reference")
    notes: List[str] = []

    if mode == "go_to_reference":
        target = deepcopy(reference_weights)
    elif mode == "absolute_targets":
        absolute_targets = active_cfg.get("absolute_target_weights_pct", {})
        if portfolio not in absolute_targets:
            raise ConfigError(
                f"active_bets.mode = 'absolute_targets', men porteføljen {portfolio} mangler i absolute_target_weights_pct."
            )
        raw_target = {category: float(value) for category, value in absolute_targets[portfolio].items()}
        target = {category: raw_target.get(category, 0.0) for category in REFERENCE_CATEGORIES}
        total = sum(target.values())
        if not math.isclose(total, 100.0, rel_tol=0, abs_tol=1e-6):
            raise ConfigError(
                f"Absolute target weights for {portfolio} summerer til {total:.6f} og ikke 100."
            )
    elif mode == "reference_plus_active_bets":
        active_by_portfolio = active_cfg.get("category_active_bets_pct", {}).get(portfolio, {})
        target = deepcopy(reference_weights)
        for category, active_bet in active_by_portfolio.items():
            if category not in REFERENCE_CATEGORIES:
                raise ConfigError(f"Ukjent kategori i active bets for {portfolio}: {category}")
            target[category] = target.get(category, 0.0) + float(active_bet)

        total = sum(target.values())
        residual = 100.0 - total
        if abs(residual) > 1e-9:
            residual_category = get_residual_category(config, portfolio)
            if residual_category not in REFERENCE_CATEGORIES:
                raise ConfigError(
                    f"Residual category '{residual_category}' for {portfolio} er ikke en gyldig kategori."
                )
            target[residual_category] = target.get(residual_category, 0.0) + residual
            notes.append(
                f"Summen av active bets var ikke null for {portfolio}; residual {residual:+.4f} %-poeng ble lagt i {residual_category}."
            )
    else:
        raise ConfigError(
            f"Ukjent active_bets.mode '{mode}'. Bruk 'go_to_reference', 'reference_plus_active_bets' eller 'absolute_targets'."
        )

    negatives = {category: value for category, value in target.items() if value < -1e-9}
    if negatives:
        raise ConfigError(f"Target for {portfolio} inneholder negative vekter: {negatives}")

    # Clean tiny negatives from floating point and re-check sum.
    target = {category: max(0.0, value) for category, value in target.items()}
    total = sum(target.values())
    if not math.isclose(total, 100.0, rel_tol=0, abs_tol=1e-6):
        raise ConfigError(f"Target weights for {portfolio} summerer til {total:.6f} og ikke 100.")
    return target, notes


def pick_trade_security(
    classified_portfolio: pd.DataFrame,
    category: str,
    config: Dict[str, Any],
    portfolio: str,
) -> str:
    routing = config.get("trade_routing", {})
    overrides = routing.get("portfolio_security_overrides", {})
    default_buy_security = routing.get("default_buy_security_by_category", {})
    prefer_existing_largest_holding = bool(routing.get("prefer_existing_largest_holding", True))

    portfolio_overrides = overrides.get(portfolio, {})
    if category in portfolio_overrides:
        return portfolio_overrides[category]

    if prefer_existing_largest_holding:
        existing = classified_portfolio[classified_portfolio["category"] == category].copy()
        if not existing.empty:
            existing = existing.sort_values(["mv", "security_name"], ascending=[False, True])
            return str(existing.iloc[0]["security_name"])

    if category in default_buy_security:
        return str(default_buy_security[category])

    raise ConfigError(
        f"Fant ikke trade vehicle for kategori '{category}' i portefølje '{portfolio}'. "
        "Legg inn trade_routing.default_buy_security_by_category eller portfolio_security_overrides."
    )


def build_trade_tables(
    classified: pd.DataFrame,
    current_weights: pd.DataFrame,
    config: Dict[str, Any],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    portfolios = sorted(classified["portfolio"].unique().tolist())
    round_decimals = int(config.get("round_decimals", 4))
    min_trade_pct = float(config.get("minimum_trade_pct", 0.09))

    summary_rows: List[Dict[str, Any]] = []
    reference_rows: List[Dict[str, Any]] = []
    target_rows: List[Dict[str, Any]] = []
    category_trade_rows: List[Dict[str, Any]] = []
    security_trade_rows: List[Dict[str, Any]] = []

    for portfolio in portfolios:
        portfolio_holdings = classified[classified["portfolio"] == portfolio].copy()
        total_mv = portfolio_total_mv(portfolio_holdings)
        current_map = current_category_map(current_weights, portfolio)
        reference_map = build_reference_weights_for_portfolio(portfolio, current_map, config)
        target_map, notes = build_target_weights(portfolio, reference_map, config)

        current_equity = (
            current_map.get("equity_norway", 0.0)
            + current_map.get("equity_global_developed", 0.0)
            + current_map.get("equity_global_em", 0.0)
        )
        reference_equity = (
            reference_map.get("equity_norway", 0.0)
            + reference_map.get("equity_global_developed", 0.0)
            + reference_map.get("equity_global_em", 0.0)
        )
        target_equity = (
            target_map.get("equity_norway", 0.0)
            + target_map.get("equity_global_developed", 0.0)
            + target_map.get("equity_global_em", 0.0)
        )
        summary_rows.append(
            {
                "portfolio": portfolio,
                "portfolio_mv": total_mv,
                "current_equity_pct": round(current_equity, round_decimals),
                "reference_equity_pct": round(reference_equity, round_decimals),
                "target_equity_pct": round(target_equity, round_decimals),
                "minimum_trade_pct": min_trade_pct,
                "notes": " | ".join(notes),
            }
        )

        for category in REFERENCE_CATEGORIES:
            current_pct = current_map.get(category, 0.0)
            reference_pct = reference_map.get(category, 0.0)
            target_pct = target_map.get(category, 0.0)
            active_bet_vs_reference_pct = current_pct - reference_pct
            trade_pct = target_pct - current_pct
            trade_mv = total_mv * trade_pct / 100.0
            skip_trade = abs(trade_pct) < min_trade_pct

            reference_rows.append(
                {
                    "portfolio": portfolio,
                    "category": category,
                    "category_name": CATEGORY_DESCRIPTIONS.get(category, category),
                    "current_weight_pct": round(current_pct, round_decimals),
                    "reference_weight_pct": round(reference_pct, round_decimals),
                    "active_bet_vs_reference_pct": round(active_bet_vs_reference_pct, round_decimals),
                }
            )
            target_rows.append(
                {
                    "portfolio": portfolio,
                    "category": category,
                    "category_name": CATEGORY_DESCRIPTIONS.get(category, category),
                    "target_weight_pct": round(target_pct, round_decimals),
                }
            )
            category_trade_rows.append(
                {
                    "portfolio": portfolio,
                    "category": category,
                    "category_name": CATEGORY_DESCRIPTIONS.get(category, category),
                    "current_weight_pct": round(current_pct, round_decimals),
                    "target_weight_pct": round(target_pct, round_decimals),
                    "trade_weight_pct": round(trade_pct, round_decimals),
                    "trade_mv": round(trade_mv, 2),
                    "skip_due_to_minimum": skip_trade,
                }
            )

            if skip_trade:
                continue

            trade_security = pick_trade_security(portfolio_holdings, category, config, portfolio)
            security_trade_rows.append(
                {
                    "portfolio": portfolio,
                    "category": category,
                    "category_name": CATEGORY_DESCRIPTIONS.get(category, category),
                    "security_name": trade_security,
                    "trade_weight_pct": round(trade_pct, round_decimals),
                    "trade_mv": round(trade_mv, 2),
                    "side": "BUY" if trade_mv > 0 else "SELL",
                }
            )

    return (
        pd.DataFrame(summary_rows),
        pd.DataFrame(reference_rows),
        pd.DataFrame(target_rows),
        pd.DataFrame(category_trade_rows),
        pd.DataFrame(security_trade_rows),
    )


def apply_formatting_to_excel(output_path: Path) -> None:
    wb = load_workbook(output_path)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        max_col = ws.max_column
        max_row = ws.max_row

        # General formatting and auto-width.
        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 38)

        # Number formatting by header names.
        headers = {str(ws.cell(1, c).value): c for c in range(1, max_col + 1) if ws.cell(1, c).value is not None}
        for header, col_idx in headers.items():
            if header.endswith("_pct") or header.endswith("_weight_pct") or header == "minimum_trade_pct":
                for r in range(2, max_row + 1):
                    ws.cell(r, col_idx).number_format = "0.0000"
            elif header.endswith("_mv") or header == "portfolio_mv":
                for r in range(2, max_row + 1):
                    ws.cell(r, col_idx).number_format = '#,##0.00'

        # Highlight skipped trades on trade sheets.
        if ws.title == "category_trades":
            headers = {str(ws.cell(1, c).value): c for c in range(1, max_col + 1) if ws.cell(1, c).value is not None}
            skip_col = headers.get("skip_due_to_minimum")
            if skip_col:
                for r in range(2, max_row + 1):
                    if ws.cell(r, skip_col).value is True:
                        for c in range(1, max_col + 1):
                            ws.cell(r, c).fill = SECTION_FILL

    wb.save(output_path)


def config_to_dataframe(config: Dict[str, Any]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    def walk(prefix: str, value: Any) -> None:
        if isinstance(value, dict):
            for key, sub_value in value.items():
                walk(f"{prefix}.{key}" if prefix else str(key), sub_value)
        else:
            rows.append({"key": prefix, "value": json.dumps(value, ensure_ascii=False) if isinstance(value, list) else value})

    walk("", config)
    return pd.DataFrame(rows)


def write_output_workbook(
    output_path: Path,
    classified: pd.DataFrame,
    current_weights: pd.DataFrame,
    summary: pd.DataFrame,
    reference_vs_active: pd.DataFrame,
    target_weights: pd.DataFrame,
    category_trades: pd.DataFrame,
    security_trades: pd.DataFrame,
    config: Dict[str, Any],
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="summary")
        reference_vs_active.to_excel(writer, index=False, sheet_name="reference_vs_active")
        target_weights.to_excel(writer, index=False, sheet_name="target_weights")
        category_trades.to_excel(writer, index=False, sheet_name="category_trades")
        security_trades.to_excel(writer, index=False, sheet_name="security_trades")
        current_weights.to_excel(writer, index=False, sheet_name="current_category_weights")
        classified.to_excel(writer, index=False, sheet_name="classified_holdings")
        config_to_dataframe(config).to_excel(writer, index=False, sheet_name="config_used")

    apply_formatting_to_excel(output_path)


def run_rebalancer(input_path: Path, config_path: Path, output_path: Path) -> None:
    config = load_json(config_path)
    exposure_column = config.get("exposure_column", "Lev. expo. distr. (PF)")
    holdings = read_workbook_holdings(input_path, exposure_column=exposure_column)
    classified = classify_holdings(holdings, config)
    current_weights = summarize_current_weights(classified)
    summary, reference_vs_active, target_weights, category_trades, security_trades = build_trade_tables(
        classified, current_weights, config
    )
    write_output_workbook(
        output_path=output_path,
        classified=classified,
        current_weights=current_weights,
        summary=summary,
        reference_vs_active=reference_vs_active,
        target_weights=target_weights,
        category_trades=category_trades,
        security_trades=security_trades,
        config=config,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Portfolio rebalancer for multi-sheet portfolio workbooks.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init-config", help="Create a starter JSON config from the workbook.")
    init_parser.add_argument("--input", required=True, type=Path, help="Input workbook (.xlsx)")
    init_parser.add_argument("--config-out", required=True, type=Path, help="Where to write the starter JSON config")
    init_parser.add_argument(
        "--exposure-column",
        default="Lev. expo. distr. (PF)",
        help="Column to use for current exposures, for example 'Lev. expo. distr. (PF)' or 'Lev. expo. distr.pre-sim.'",
    )

    run_parser = subparsers.add_parser("run", help="Run the rebalance and write an output workbook.")
    run_parser.add_argument("--input", required=True, type=Path, help="Input workbook (.xlsx)")
    run_parser.add_argument("--config", required=True, type=Path, help="JSON config file")
    run_parser.add_argument("--output", required=True, type=Path, help="Output workbook (.xlsx)")

    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        if args.command == "init-config":
            config = make_default_config(args.input, exposure_column=args.exposure_column)
            write_json(args.config_out, config)
            print(f"Starter config written to {args.config_out}")
            return 0

        if args.command == "run":
            run_rebalancer(args.input, args.config, args.output)
            print(f"Rebalance workbook written to {args.output}")
            return 0

        parser.error("Unknown command")
        return 2

    except (ConfigError, DataError, FileNotFoundError, json.JSONDecodeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
