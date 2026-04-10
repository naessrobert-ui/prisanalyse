"""
portfolio_rebalancer_engine.py
------------------------------
Bridge module that exposes the API expected by portfolio_rebalancer_routes.py
while delegating all business logic to the original portfolio_rebalancer.py.

portfolio_rebalancer.py must be importable (same directory or on sys.path).

Routes expect these public functions:
    workbook_to_frames(raw_bytes)          -> dict[str, DataFrame]
    detect_portfolio_sheets(frames)        -> list[str]
    extract_holdings(frames, sheets)       -> DataFrame
    classify_holdings(holdings, mapping)   -> DataFrame
    compute_actual_exposures(classified)   -> DataFrame
    compute_benchmark_exposures(df)        -> DataFrame
    compute_target_exposures(...)          -> DataFrame
    compute_active_bets(...)               -> DataFrame
    export_results_to_excel(frames)        -> bytes
"""

from __future__ import annotations

import io
import math
import re
from copy import deepcopy
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook as openpyxl_load_workbook

# ---------------------------------------------------------------------------
# Import the original rebalancer module (must live alongside this file).
# ---------------------------------------------------------------------------
import portfolio_rebalancer as _rb


# ---------------------------------------------------------------------------
# 1. Workbook I/O helpers
# ---------------------------------------------------------------------------

def workbook_to_frames(raw_bytes: bytes) -> Dict[str, pd.DataFrame]:
    """
    Read an Excel workbook from raw bytes and return a dict of
    {sheet_name: DataFrame} with raw cell values (data_only=True).

    The DataFrames are indexed from row 0 (header row) onward so that
    downstream functions can inspect them, but extract_holdings() will
    use the original openpyxl row-iteration logic.
    """
    wb = openpyxl_load_workbook(io.BytesIO(raw_bytes), data_only=True)
    frames: Dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            frames[sheet_name] = pd.DataFrame()
            continue
        headers = [_rb.normalize_text(v) for v in rows[0]]
        data = [list(row) for row in rows[1:]]
        frames[sheet_name] = pd.DataFrame(data, columns=headers)
    return frames


def detect_portfolio_sheets(frames: Dict[str, pd.DataFrame]) -> List[str]:
    """
    Return sheet names that look like portfolio sheets, i.e. they contain
    the expected columns. Sheets missing required columns are skipped.
    """
    required = {"portfolio", "security name", "mv"}
    detected = []
    for name, df in frames.items():
        cols = {c.lower() for c in df.columns if isinstance(c, str)}
        if required.issubset(cols):
            detected.append(name)
    return detected


# ---------------------------------------------------------------------------
# 2. Holdings extraction
# ---------------------------------------------------------------------------

def extract_holdings(
    frames: Dict[str, pd.DataFrame],
    sheets: List[str],
    exposure_column: str = "Lev. expo. distr. (PF)",
) -> pd.DataFrame:
    """
    Re-implement the row-level extraction logic from read_workbook_holdings()
    but operating on already-loaded DataFrames (so we don't need a file path).
    """
    exposure_col_norm = _rb.normalize_text(exposure_column)
    records: List[Dict[str, Any]] = []

    for sheet_name in sheets:
        df = frames.get(sheet_name)
        if df is None or df.empty:
            continue

        # Build a case-insensitive header map  col_normalized -> col_actual
        col_map: Dict[str, str] = {_rb.normalize_text(c): c for c in df.columns}

        required = ["portfolio", "security name", "mv", exposure_col_norm]
        missing = [r for r in required if r not in col_map]
        if missing:
            raise _rb.DataError(
                f"Fanen '{sheet_name}' mangler kolonner: {missing}. "
                f"Fant: {list(col_map.keys())}"
            )

        portfolio_col = col_map["portfolio"]
        security_col = col_map["security name"]
        mv_col = col_map["mv"]
        expo_col = col_map[exposure_col_norm]
        model_col = col_map.get("model portfolio")
        country_col = col_map.get("country")

        for row_number, (_, row) in enumerate(df.iterrows(), start=2):
            portfolio = _rb.normalize_text(row[portfolio_col])
            security = row[security_col]
            mv = _rb.safe_float(row[mv_col])
            expo = row[expo_col]
            model_portfolio = _rb.normalize_text(row[model_col]) if model_col else ""
            country = _rb.normalize_text(row[country_col]) if country_col else ""

            if expo is None:
                continue

            security_name = _rb.normalize_text(security)
            if not security_name:
                continue

            is_cash_position = security_name.startswith("PB Norway:")
            is_security_holding = bool(model_portfolio)
            if not (is_cash_position or is_security_holding):
                continue

            records.append(
                {
                    "sheet_name": sheet_name,
                    "portfolio": sheet_name,
                    "internal_portfolio_name": portfolio,
                    "row_number": row_number,
                    "security_name": security_name,
                    "country": country,
                    "model_portfolio": model_portfolio,
                    "mv": mv,
                    "exposure_pct": _rb.safe_float(expo),
                }
            )

    if not records:
        raise _rb.DataError("Fant ingen beholdningsrader i arbeidsboken.")

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# 3. Classification
# ---------------------------------------------------------------------------

def classify_holdings(
    holdings: pd.DataFrame,
    mapping: Dict[str, str],
) -> pd.DataFrame:
    """
    Classify holdings using a flat {security_name: category} mapping dict.
    This is the mapping_json from the UI (not a full config dict).
    """
    # Build a minimal config so we can reuse the original function.
    config = {"security_categories": mapping}
    return _rb.classify_holdings(holdings, config)


# ---------------------------------------------------------------------------
# 4. Exposure computations
# ---------------------------------------------------------------------------

def compute_actual_exposures(classified: pd.DataFrame) -> pd.DataFrame:
    """Wraps summarize_current_weights."""
    return _rb.summarize_current_weights(classified)


def compute_benchmark_exposures(benchmark_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert the benchmark rows DataFrame (from benchmark_json) into a format
    that mirrors current_weights so we can pass it through the reference logic.

    Input columns expected:
        portfolio, equity_share, norway_share_within_equity,
        em_share_within_international_equity

    Output columns:
        portfolio, category, reference_weight_pct
    """
    rows: List[Dict[str, Any]] = []

    for _, r in benchmark_df.iterrows():
        portfolio = str(r["portfolio"])
        equity_share = float(r.get("equity_share", 0.60))
        norway_share = float(r.get("norway_share_within_equity", 0.20))
        em_share = float(r.get("em_share_within_international_equity", 0.15))

        equity_total = equity_share * 100.0
        equity_norway = equity_total * norway_share
        international_equity = equity_total - equity_norway
        equity_em = international_equity * em_share
        equity_dm = international_equity - equity_em
        fi_total = 100.0 - equity_total

        # Default: split fi 50/50 Norway/Global, Norway 50/50 short/long
        fi_norway = fi_total * 0.5
        fi_short = fi_norway * 0.5
        fi_long = fi_norway * 0.5
        fi_global = fi_total - fi_norway

        rows.extend(
            [
                {"portfolio": portfolio, "category": "cash", "reference_weight_pct": 0.0},
                {"portfolio": portfolio, "category": "allocation", "reference_weight_pct": 0.0},
                {"portfolio": portfolio, "category": "equity_norway", "reference_weight_pct": equity_norway},
                {"portfolio": portfolio, "category": "equity_global_developed", "reference_weight_pct": equity_dm},
                {"portfolio": portfolio, "category": "equity_global_em", "reference_weight_pct": equity_em},
                {"portfolio": portfolio, "category": "fi_norway_short", "reference_weight_pct": fi_short},
                {"portfolio": portfolio, "category": "fi_norway_long", "reference_weight_pct": fi_long},
                {"portfolio": portfolio, "category": "fi_global", "reference_weight_pct": fi_global},
            ]
        )

    return pd.DataFrame(rows)


def compute_target_exposures(
    benchmark_exposures: pd.DataFrame,
    rebalance_mode: str = "go_to_benchmark",
    active_bets: Optional[pd.DataFrame] = None,
    absolute_targets: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    Compute per-portfolio target weights.

    Modes:
        go_to_benchmark       -> target = benchmark
        reference_plus_active -> target = benchmark + active_bets (with residual in fi_global)
        absolute_targets      -> target = absolute_targets

    Output columns: portfolio, category, target_weight_pct
    """
    rows: List[Dict[str, Any]] = []

    portfolios = benchmark_exposures["portfolio"].unique().tolist()

    for portfolio in portfolios:
        ref = (
            benchmark_exposures[benchmark_exposures["portfolio"] == portfolio]
            .set_index("category")["reference_weight_pct"]
            .to_dict()
        )

        if rebalance_mode in ("go_to_benchmark", "go_to_reference"):
            target = deepcopy(ref)

        elif rebalance_mode == "reference_plus_active" and active_bets is not None and not active_bets.empty:
            target = deepcopy(ref)
            port_bets = active_bets[active_bets["portfolio"] == portfolio] if "portfolio" in active_bets.columns else active_bets
            for _, bet_row in port_bets.iterrows():
                cat = str(bet_row.get("category", ""))
                val = float(bet_row.get("active_bet_pct", 0.0))
                if cat in target:
                    target[cat] = target.get(cat, 0.0) + val
            # Residual absorbed by fi_global
            total = sum(target.values())
            residual = 100.0 - total
            if abs(residual) > 1e-9:
                target["fi_global"] = target.get("fi_global", 0.0) + residual

        elif rebalance_mode == "absolute_targets" and absolute_targets is not None and not absolute_targets.empty:
            port_abs = (
                absolute_targets[absolute_targets["portfolio"] == portfolio]
                if "portfolio" in absolute_targets.columns
                else absolute_targets
            )
            target = {str(r["category"]): float(r["target_weight_pct"]) for _, r in port_abs.iterrows()}

        else:
            target = deepcopy(ref)

        for cat, weight in target.items():
            rows.append({"portfolio": portfolio, "category": cat, "target_weight_pct": weight})

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 5. Active bets / trade table
# ---------------------------------------------------------------------------

def compute_active_bets(
    actual: pd.DataFrame,
    benchmark: pd.DataFrame,
    target: pd.DataFrame,
    minimum_trade_threshold: float = 0.01,
) -> pd.DataFrame:
    """
    Merge actual, benchmark, and target exposures into a single trade table.

    Output columns:
        portfolio, category, current_weight_pct, reference_weight_pct,
        target_weight_pct, active_bet_vs_reference_pct,
        trade_weight_pct, skip_due_to_minimum
    """
    # Merge on portfolio + category
    merged = (
        actual.rename(columns={"current_weight_pct": "current_weight_pct"})
        .merge(
            benchmark[["portfolio", "category", "reference_weight_pct"]],
            on=["portfolio", "category"],
            how="outer",
        )
        .merge(
            target[["portfolio", "category", "target_weight_pct"]],
            on=["portfolio", "category"],
            how="outer",
        )
    )

    merged["current_weight_pct"] = merged["current_weight_pct"].fillna(0.0)
    merged["reference_weight_pct"] = merged["reference_weight_pct"].fillna(0.0)
    merged["target_weight_pct"] = merged["target_weight_pct"].fillna(0.0)

    merged["active_bet_vs_reference_pct"] = merged["current_weight_pct"] - merged["reference_weight_pct"]
    merged["trade_weight_pct"] = merged["target_weight_pct"] - merged["current_weight_pct"]
    merged["skip_due_to_minimum"] = merged["trade_weight_pct"].abs() < minimum_trade_threshold

    return merged.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 6. Excel export
# ---------------------------------------------------------------------------

def export_results_to_excel(frames: Dict[str, pd.DataFrame]) -> bytes:
    """
    Write all result DataFrames to an Excel workbook in memory and return bytes.
    Sheet order follows the dict insertion order.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])  # Excel 31-char limit
    buf.seek(0)
    return buf.read()
